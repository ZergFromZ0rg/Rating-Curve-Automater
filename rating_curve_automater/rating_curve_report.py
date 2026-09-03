from __future__ import annotations

from pathlib import Path
from typing import Callable

import numpy as np
import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill

from rating_curve_automater.rating_curve_drift import assess_temporal_drift
from rating_curve_automater.rating_curve_fitting import select_valid_measurements
from rating_curve_automater.rating_table import DEFAULT_STAGE_STEP_M, build_rating_table
from rating_curve_automater.schema import DATE, DISCHARGE_CMS, SITE, STAGE_M

# Friendly column labels used in the exported workbook.
OUT_DATE = "Date"
OUT_STAGE = "Stage Above Bed (m)"
OUT_OBSERVED = "Measured Discharge Q (m³/s)"
OUT_MODELED = "Modeled Discharge Q (m³/s)"


def build_observed_modeled_table(
    df: pd.DataFrame,
    a: float,
    b: float,
    h0: float,
    uncertainty_threshold: float = 0.25,
    predict: Callable[[np.ndarray], np.ndarray] | None = None,
) -> pd.DataFrame:
    """Build an observed-vs-modeled table with uncertainty flags.

    Operates on the same valid rows as :func:`fit_rating_curve` so the report
    reflects the data the model was actually fitted on. ``predict`` overrides
    the single power law (used for segmented curves).
    """
    working = select_valid_measurements(df)

    stage = working[STAGE_M].astype(float).to_numpy()
    observed = working[DISCHARGE_CMS].astype(float).to_numpy()
    if predict is not None:
        modeled = np.asarray(predict(stage), dtype=float)
    else:
        modeled = a * np.power(np.maximum(stage - h0, 1e-9), b)
    residual = observed - modeled
    relative_error = np.abs(residual / np.maximum(observed, 1e-9))

    flags = np.where(relative_error > uncertainty_threshold, "Uncertain", "Normal")

    output = pd.DataFrame({
        OUT_DATE: working.get(DATE, pd.Series(index=working.index, dtype="object")),
        OUT_STAGE: stage,
        OUT_OBSERVED: observed,
        OUT_MODELED: modeled,
        "Residual": residual,
        "Relative Error": relative_error,
        "Uncertainty Flag": flags,
    })

    return output


def build_summary_table(fit: dict, table: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if fit.get("site"):
        rows.append({"Metric": "site", "Value": fit["site"]})
    rows.append({
        "Metric": "estimator",
        "Value": "Bayesian (ratingcurve / PyMC)" if fit.get("method") == "bayesian" else "log-log least squares",
    })
    rows.append({"Metric": "h0", "Value": fit["h0"]})

    if fit.get("is_segmented"):
        breakpoints = fit.get("breakpoints", [fit.get("breakpoint")])
        pick = f" (chosen by {fit.get('criterion', 'bic').upper()})" if fit.get("segment_selection") == "auto" else ""
        rows.append({"Metric": "segments", "Value": f"{fit.get('n_segments', len(breakpoints) + 1)}{pick}"})
        rows.append({"Metric": "breakpoint stage(s) (m)", "Value": ", ".join(f"{bp:.3f}" for bp in breakpoints)})
        bp_ci = fit.get("breakpoint_ci")
        if bp_ci:
            lvl = int(round(fit.get("bands", {}).get("level", 0.95) * 100))
            rows.append({"Metric": f"breakpoint {lvl}% credible interval(s) (m)",
                         "Value": "; ".join(f"[{lo:.3f}, {hi:.3f}]" for lo, hi in bp_ci)})
        for i, seg in enumerate(fit["segments"], start=1):
            rows.append({"Metric": f"segment {i} a", "Value": seg["a"]})
            rows.append({"Metric": f"segment {i} b", "Value": seg["b"]})
            rows.append({"Metric": f"segment {i} points", "Value": seg["n_points"]})
    else:
        rows.append({"Metric": "a", "Value": fit["a"]})
        rows.append({"Metric": "b", "Value": fit["b"]})
        if fit.get("b_fixed"):
            rows.append({"Metric": "b source", "Value": "imposed by user (not fitted)"})

    if fit.get("uncertainty_source") == "column":
        kind = "weighted least squares" if fit.get("weighted") else "uniform (no re-weighting)"
        rows.append({"Metric": "discharge uncertainty", "Value": f"per-point column, {kind}"})
        if fit.get("mean_uncertainty_pct") is not None:
            rows.append({"Metric": "mean discharge uncertainty (%)", "Value": round(fit["mean_uncertainty_pct"], 2)})
    elif fit.get("uncertainty_pct_default") is not None:
        rows.append({"Metric": "discharge uncertainty", "Value": f"{fit['uncertainty_pct_default']:.1f}% assumed (no re-weighting)"})

    rows += [
        {"Metric": "R^2", "Value": fit["r_squared"]},
    ]
    if fit.get("weighted") and fit.get("r_squared_weighted") is not None:
        rows.append({"Metric": "weighted R^2", "Value": fit["r_squared_weighted"]})

    bands = fit.get("bands")
    if bands:
        pct = int(round(bands["level"] * 100))
        if bands.get("b_ci"):
            lo, hi = bands["b_ci"]
            rows.append({"Metric": f"{pct}% CI on b", "Value": f"[{lo:.4f}, {hi:.4f}]"})
        if bands.get("a_ci"):
            lo, hi = bands["a_ci"]
            rows.append({"Metric": f"{pct}% CI on a", "Value": f"[{lo:.4f}, {hi:.4f}]"})
        if bands.get("h0_ci"):
            lo, hi = bands["h0_ci"]
            src = "posterior" if bands.get("kind") == "posterior" else "re-estimated per replicate"
            rows.append({"Metric": f"{pct}% CI on h0 ({src})",
                         "Value": f"[{lo:.3f}, {hi:.3f}]"})
        rows.append({
            "Metric": f"{pct}% confidence band half-width at median stage (%)",
            "Value": round(bands["ci_halfwidth_pct_at_median"], 1),
        })
        if bands.get("kind") == "posterior":
            rows.append({"Metric": "posterior draws", "Value": bands["n_success"]})
        else:
            rows.append({
                "Metric": "bootstrap replicates (succeeded / requested)",
                "Value": f"{bands['n_success']} / {bands['n_bootstrap']}",
            })

    rows += [
        {"Metric": "Valid points", "Value": len(table)},
        {"Metric": "Uncertain points", "Value": int((table["Uncertainty Flag"] == "Uncertain").sum())},
        {"Metric": "Normal points", "Value": int((table["Uncertainty Flag"] == "Normal").sum())},
    ]

    loo = fit.get("loo")
    if loo:
        rows += [
            {"Metric": "leave-one-out RMSPE (%)", "Value": round(loo["rmspe_pct"], 2)},
            {"Metric": "leave-one-out bias (%)", "Value": round(loo["bias_pct"], 2)},
            {"Metric": "leave-one-out 95th-pct abs error (%)", "Value": round(loo["p95_abs_pct"], 1)},
        ]

    manning = fit.get("manning")
    if manning and manning.get("flag") not in (None, "unusable"):
        rows.append({"Metric": "Manning cross-section check", "Value": manning["flag"]})
        if manning.get("n_used") is not None:
            rows.append({"Metric": "Manning's n (calibrated to rating)" if manning.get("n_supplied") is None
                         else "Manning's n (supplied)", "Value": round(manning["n_used"], 4)})
        if np.isfinite(manning.get("max_abs_pct_diff_extrapolated", float("nan"))):
            rows.append({"Metric": "fitted-vs-Manning max diff, extrapolated (%)",
                         "Value": round(manning["max_abs_pct_diff_extrapolated"], 1)})
        rows.append({"Metric": "Manning check note", "Value": manning["message"]})
    drift = fit.get("drift")
    if drift:
        rows.append({"Metric": "temporal drift flag", "Value": drift["flag"]})
        if drift.get("trend_pct_per_year") is not None:
            rows.append({"Metric": "residual trend (%/yr)", "Value": round(drift["trend_pct_per_year"], 2)})
            rows.append({"Metric": "residual trend p-value", "Value": round(drift["trend_p_value"], 3)})
            rows.append({"Metric": f"recent {drift['recent_n']} gaugings mean offset (%)",
                         "Value": round(drift["recent_mean_pct"], 1)})
        cp = drift.get("changepoint")
        if cp is not None:
            rows.append({"Metric": "drift changepoint date", "Value": cp["date"]})
            rows.append({"Metric": "changepoint shift (%)", "Value": round(cp["shift_pct"], 1)})
            rows.append({"Metric": "changepoint p-value", "Value": round(cp["p_value"], 3)})
            rows.append({"Metric": "gaugings before / after changepoint",
                         "Value": f"{cp['n_before']} / {cp['n_after']}"})
        rows.append({"Metric": "temporal drift note", "Value": drift["message"]})

    if not fit.get("is_plausible", True):
        rows.append({"Metric": "PLAUSIBILITY", "Value": "FAILED - see warnings"})
    for i, warning in enumerate(fit.get("warnings", []), start=1):
        rows.append({"Metric": f"warning {i}", "Value": warning})
    return pd.DataFrame(rows)


def export_rating_curve_report(
    df: pd.DataFrame,
    output_path: str | Path,
    a: float,
    b: float,
    h0: float,
    uncertainty_threshold: float = 0.25,
    r_squared: float | None = None,
    fit: dict | None = None,
    site: str | None = None,
    rating_table_step: float = DEFAULT_STAGE_STEP_M,
) -> Path:
    """Write the multi-sheet Excel report.

    Pass ``fit`` (the dict from :func:`fit_rating_curve`) to render segmented
    curves and reuse its overall R²; otherwise a single power law ``a``/``b``/
    ``h0`` is used. ``site`` (or a single-valued site column in ``df``) is
    recorded in the Summary sheet. ``rating_table_step`` is the stage increment
    (m) for the *Rating Table* sheet.
    """
    from rating_curve_automater.rating_curve_fitting import predict_discharge

    predict = (lambda stage: predict_discharge(fit, stage)) if fit is not None else None

    if site is None and SITE in df.columns:
        unique_sites = df[SITE].dropna().astype(str).str.strip().unique()
        unique_sites = [s for s in unique_sites if s]
        if len(unique_sites) == 1:
            site = unique_sites[0]
    original_data = df.copy()
    table = build_observed_modeled_table(
        df, a=a, b=b, h0=h0, uncertainty_threshold=uncertainty_threshold, predict=predict
    )

    if fit is not None and r_squared is None:
        r_squared = fit.get("r_squared")

    if r_squared is None:
        if len(table) > 1:
            ss_res = float(np.sum((table[OUT_OBSERVED] - table[OUT_MODELED]) ** 2))
            ss_tot = float(np.sum((table[OUT_OBSERVED] - table[OUT_OBSERVED].mean()) ** 2))
            r_squared = 1.0 - (ss_res / ss_tot) if ss_tot != 0 else 1.0
        else:
            r_squared = 1.0

    rating_fit = dict(fit) if fit is not None else {"a": a, "b": b, "h0": h0, "is_segmented": False}
    rating_fit.setdefault("stage_min", float(table[OUT_STAGE].min()))
    rating_fit.setdefault("stage_max", float(table[OUT_STAGE].max()))

    drift = rating_fit.get("drift")
    if drift is None:
        drift = assess_temporal_drift(select_valid_measurements(df), rating_fit, random_state=0)

    fit_summary = dict(fit) if fit is not None else {"a": a, "b": b, "h0": h0}
    fit_summary["r_squared"] = r_squared
    if drift is not None:
        fit_summary["drift"] = drift
    if fit_summary.get("loo") is None:
        from rating_curve_automater.rating_curve_fitting import leave_one_out_error

        fdict = fit or {}
        if fdict.get("segment_selection") == "auto":
            segments_arg: int | str = "auto"
        else:
            segments_arg = fdict.get("n_segments", 1) if fdict.get("n_segments", 1) > 1 else 1
        try:
            fit_summary["loo"] = leave_one_out_error(
                df, segments=segments_arg,
                h0=None if fdict.get("h0_estimated", True) else h0,
                discharge_uncertainty_pct=fdict.get("uncertainty_pct_default", 7.0),
                fixed_b=b if fdict.get("b_fixed") else None,
            )
        except Exception:  # noqa: BLE001 - a diagnostic must never break the report
            fit_summary["loo"] = None
    if site:
        fit_summary["site"] = site
    summary = build_summary_table(fit_summary, table)
    try:
        rating_table = build_rating_table(rating_fit, step=rating_table_step)
    except ValueError:
        rating_table = None
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        original_data.to_excel(writer, sheet_name="Original Data", index=False)
        table.to_excel(writer, sheet_name="Observed vs Modeled", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        if rating_table is not None and not rating_table.empty:
            rating_table.to_excel(writer, sheet_name="Rating Table", index=False)

        plot_data = pd.DataFrame(
            {
                "Stage Above Bed (m)": table[OUT_STAGE],
                "Observed Discharge Q (m³/s)": table[OUT_OBSERVED],
                "Modeled Discharge Q (m³/s)": table[OUT_MODELED],
                "Uncertain Observed Discharge Q (m³/s)": np.where(table["Uncertainty Flag"] == "Uncertain", table[OUT_OBSERVED], np.nan),
            }
        )
        plot_data.to_excel(writer, sheet_name="Plot Data", index=False)

        workbook = writer.book
        plot_ws = workbook.create_sheet("Plot")
        plot_ws.append(["Stage Above Bed (m)", "Observed Discharge Q (m³/s)", "Modeled Discharge Q (m³/s)", "Uncertain Observed Discharge Q (m³/s)"])
        for row in plot_data.values.tolist():
            plot_ws.append(row)

        chart = LineChart()
        chart.title = "Rating Curve"
        chart.x_axis.title = "Stage Above Bed (m)"
        chart.y_axis.title = "Discharge (m³/s)"
        chart.style = 13
        chart.legend.position = "r"
        chart.height = 9
        chart.width = 18
        chart.y_axis.scaling.min = 0

        observed_data = Reference(plot_ws, min_col=2, max_col=2, min_row=1, max_row=plot_ws.max_row)
        modeled_data = Reference(plot_ws, min_col=3, max_col=3, min_row=1, max_row=plot_ws.max_row)
        uncertain_data = Reference(plot_ws, min_col=4, max_col=4, min_row=1, max_row=plot_ws.max_row)
        categories = Reference(plot_ws, min_col=1, max_col=1, min_row=2, max_row=plot_ws.max_row)

        chart.add_data(observed_data, titles_from_data=True)
        chart.add_data(modeled_data, titles_from_data=True)
        chart.add_data(uncertain_data, titles_from_data=True)
        chart.set_categories(categories)

        chart.series[0].graphicalProperties.line.solidFill = "1F77B4"
        chart.series[0].graphicalProperties.line.width = 2
        chart.series[0].marker.size = 6
        chart.series[0].marker.symbol = "circle"
        chart.series[0].marker.graphicalProperties.solidFill = "1F77B4"

        chart.series[1].graphicalProperties.line.solidFill = "2CA02C"
        chart.series[1].graphicalProperties.line.width = 2
        chart.series[1].marker.size = 4
        chart.series[1].marker.symbol = "triangle"
        chart.series[1].marker.graphicalProperties.solidFill = "2CA02C"

        chart.series[2].graphicalProperties.line.noFill = True
        chart.series[2].marker.size = 7
        chart.series[2].marker.symbol = "diamond"
        chart.series[2].marker.graphicalProperties.solidFill = "D62728"

        plot_ws.add_chart(chart, "F2")

        observed_ws = workbook["Observed vs Modeled"]
        flagged_fill = PatternFill(fill_type="solid", fgColor="FFC000")
        normal_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        for row_idx in range(2, observed_ws.max_row + 1):
            flag = observed_ws.cell(row=row_idx, column=7).value
            fill = flagged_fill if flag == "Uncertain" else normal_fill
            for col_idx in range(1, observed_ws.max_column + 1):
                observed_ws.cell(row=row_idx, column=col_idx).fill = fill

        bands = fit.get("bands") if fit is not None else None
        if bands:
            _write_band_sheet(writer, bands)

        if drift and isinstance(drift.get("residuals"), pd.DataFrame) and not drift["residuals"].empty:
            _write_residuals_over_time_sheet(writer, drift)

        manning = fit.get("manning") if fit is not None else None
        if manning and "stage" in manning:
            _write_manning_sheet(writer, manning)

    return output


def _write_manning_sheet(writer, manning: dict) -> None:
    """'Manning Check' sheet: the fitted curve vs a Manning curve from the
    surveyed cross-section, over the gauged range and the extrapolation."""
    df = pd.DataFrame({
        "Stage (m)": np.round(np.asarray(manning["stage"]), 4),
        "Fitted Q (m³/s)": np.round(np.asarray(manning["q_rating"]), 4),
        "Manning Q (m³/s)": np.round(np.asarray(manning["q_manning"]), 4),
        "Difference (%)": np.round(np.asarray(manning["pct_diff"]), 1),
        "Above highest gauging": np.asarray(manning["stage"]) > manning["stage_max_gauged"],
    })
    df.to_excel(writer, sheet_name="Manning Check", index=False)
    ws = writer.book["Manning Check"]
    ws["G2"] = "flag"
    ws["H2"] = manning["flag"]
    ws["G3"] = "Manning's n used"
    ws["H3"] = round(manning.get("n_used", float("nan")), 4)
    ws["G4"] = "channel slope"
    ws["H4"] = manning.get("slope")
    ws["G5"] = "message"
    ws["H5"] = manning["message"]

    chart = LineChart()
    chart.title = f"Fitted curve vs cross-section (Manning) — {manning['flag']}"
    chart.x_axis.title = "Stage (m)"
    chart.y_axis.title = "Discharge (m³/s)"
    chart.style = 13
    chart.height = 9
    chart.width = 18
    for col in (2, 3):
        chart.add_data(Reference(ws, min_col=col, max_col=col, min_row=1, max_row=ws.max_row),
                       titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row))
    ws.add_chart(chart, "G8")


def _write_residuals_over_time_sheet(writer, drift: dict) -> None:
    """'Residuals Over Time' sheet: each gauging's percent difference from the
    curve against its date, plus a chart."""
    resid = drift["residuals"].copy()
    resid.to_excel(writer, sheet_name="Residuals Over Time", index=False)
    ws = writer.book["Residuals Over Time"]

    pct_col = list(resid.columns).index("Residual (%)") + 1
    chart = LineChart()
    chart.title = f"Rating-curve residuals over time — drift flag: {drift['flag']}"
    chart.x_axis.title = "Gauging date"
    chart.y_axis.title = "Observed − modelled (%)"
    chart.style = 13
    chart.height = 9
    chart.width = 18
    chart.add_data(Reference(ws, min_col=pct_col, max_col=pct_col, min_row=1, max_row=ws.max_row),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row))
    s = chart.series[0]
    s.graphicalProperties.line.noFill = True
    s.marker.symbol = "circle"
    s.marker.size = 6
    s.marker.graphicalProperties.solidFill = "1F77B4"
    ws.add_chart(chart, "J2")


def _write_band_sheet(writer, bands: dict) -> None:
    """Add a 'Rating Curve Band' sheet: dense stage grid with the fitted curve
    and its confidence / prediction envelopes, plus a line chart."""
    pct = int(round(bands["level"] * 100))
    band_df = pd.DataFrame({
        "Stage Above Bed (m)": np.asarray(bands["stage"], dtype=float),
        "Modeled Q (m³/s)": np.asarray(bands["q"], dtype=float),
        f"{pct}% confidence lower": np.asarray(bands["ci_lower"], dtype=float),
        f"{pct}% confidence upper": np.asarray(bands["ci_upper"], dtype=float),
        f"{pct}% prediction lower": np.asarray(bands["pi_lower"], dtype=float),
        f"{pct}% prediction upper": np.asarray(bands["pi_upper"], dtype=float),
    })
    band_df.to_excel(writer, sheet_name="Rating Curve Band", index=False)

    ws = writer.book["Rating Curve Band"]
    chart = LineChart()
    chart.title = f"Rating curve with {pct}% confidence / prediction bands"
    chart.x_axis.title = "Stage Above Bed (m)"
    chart.y_axis.title = "Discharge (m³/s)"
    chart.style = 13
    chart.legend.position = "r"
    chart.height = 9
    chart.width = 18
    chart.y_axis.scaling.min = 0

    for col in range(2, 7):
        chart.add_data(Reference(ws, min_col=col, max_col=col, min_row=1, max_row=ws.max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row))

    chart.series[0].graphicalProperties.line.solidFill = "D62728"
    chart.series[0].graphicalProperties.line.width = 2.5
    for i in (1, 2):
        chart.series[i].graphicalProperties.line.solidFill = "1F77B4"
    for i in (3, 4):
        chart.series[i].graphicalProperties.line.solidFill = "9467BD"
    ws.add_chart(chart, "H2")


def main() -> None:
    import argparse

    import pandas as pd

    from rating_curve_automater.rating_curve_fitting import DEFAULT_N_BOOTSTRAP, fit_rating_curve
    from rating_curve_automater.rating_table import export_rating_table_csv

    parser = argparse.ArgumentParser(description="Fit a rating curve and write the Excel report.")
    parser.add_argument("--csv", type=str, default="cleaned_measurements.csv",
                        help="Cleaned measurements CSV (default: ./cleaned_measurements.csv).")
    parser.add_argument("--output", type=str, default="rating_curve_report.xlsx",
                        help="Excel report path (default: ./rating_curve_report.xlsx).")
    parser.add_argument("--step", type=float, default=DEFAULT_STAGE_STEP_M, help="Rating-table stage increment (m).")
    parser.add_argument("--rating-table-csv", type=str, default=None, help="Also write the stage-Q rating table to this CSV.")
    parser.add_argument("--bootstrap", type=int, default=DEFAULT_N_BOOTSTRAP)
    parser.add_argument("--seed", type=int, default=0)
    parser.add_argument("--segments", default="1")
    parser.add_argument("--method", choices=("ols", "bayesian"), default="ols")
    parser.add_argument(
        "--exponent", type=float, default=None, dest="fixed_b", metavar="B",
        help="Impose the power-law exponent b (e.g. 2.0) and fit only a; single power law only.",
    )
    parser.add_argument("--cross-section", type=str, default=None,
                        help="Cross-section CSV (offset + elevation) for a Manning check → a 'Manning Check' sheet.")
    parser.add_argument("--slope", type=float, default=None, help="Channel slope (m/m) — required with --cross-section.")
    parser.add_argument("--mannings-n", type=float, default=None,
                        help="Manning's n (default: calibrate it to the rating over the gauged range).")
    parser.add_argument("--stage-offset", type=float, default=0.0,
                        help="Add this to stage H to get water-surface elevation in the section's datum.")
    args = parser.parse_args()
    if args.cross_section and args.slope is None:
        raise SystemExit("--cross-section needs --slope (channel slope in m/m).")

    df = pd.read_csv(args.csv)
    segments = args.segments if args.segments.lower() == "auto" else int(args.segments)
    try:
        fit = fit_rating_curve(df, segments=segments, method=args.method,
                               n_bootstrap=args.bootstrap, random_state=args.seed,
                               fixed_b=args.fixed_b)
    except ValueError as exc:
        raise SystemExit(str(exc))
    if args.cross_section:
        from rating_curve_automater.manning import manning_sanity_check, read_cross_section

        offset, bed = read_cross_section(args.cross_section)
        fit["manning"] = manning_sanity_check(
            fit, offset, bed, args.slope, n=args.mannings_n, stage_offset=args.stage_offset,
        )
        print(f"Manning check [{fit['manning']['flag']}]: {fit['manning']['message']}")
    export_rating_curve_report(
        df, args.output, a=fit["a"], b=fit["b"], h0=fit["h0"],
        r_squared=fit["r_squared"], fit=fit, rating_table_step=args.step,
    )
    print(f"Rating-curve report written to: {Path(args.output).name}")

    if args.rating_table_csv:
        out = export_rating_table_csv(fit, args.rating_table_csv, step=args.step)
        print(f"Rating table written to: {out.name}")


if __name__ == "__main__":
    main()
