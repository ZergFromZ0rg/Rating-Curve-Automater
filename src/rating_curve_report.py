from __future__ import annotations

from pathlib import Path
from typing import Callable

import numpy as np
import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill

from src.rating_curve_fitting import select_valid_measurements
from src.schema import DATE, DISCHARGE_CMS, SITE, STAGE_M

# Friendly column labels used in the exported workbook.
OUT_DATE = "Date"
OUT_STAGE = "Stage Above Bed (m)"
OUT_OBSERVED = "Measured Discharge Q (m³/s)"
OUT_MODELED = "Modeled Discharge Q (m³/s)"


def build_observed_modeled_table(
    df: pd.DataFrame,
    a: float,
    b: float,
    h0: float,
    uncertainty_threshold: float = 0.25,
    predict: Callable[[np.ndarray], np.ndarray] | None = None,
) -> pd.DataFrame:
    """Build an observed-vs-modeled table with uncertainty flags.

    Operates on the same valid rows as :func:`fit_rating_curve` so the report
    reflects the data the model was actually fitted on. ``predict`` overrides
    the single power law (used for segmented curves).
    """
    working = select_valid_measurements(df)

    stage = working[STAGE_M].astype(float).to_numpy()
    observed = working[DISCHARGE_CMS].astype(float).to_numpy()
    if predict is not None:
        modeled = np.asarray(predict(stage), dtype=float)
    else:
        modeled = a * np.power(np.maximum(stage - h0, 1e-9), b)
    residual = observed - modeled
    relative_error = np.abs(residual / np.maximum(observed, 1e-9))

    flags = np.where(relative_error > uncertainty_threshold, "Uncertain", "Normal")

    output = pd.DataFrame({
        OUT_DATE: working.get(DATE, pd.Series(index=working.index, dtype="object")),
        OUT_STAGE: stage,
        OUT_OBSERVED: observed,
        OUT_MODELED: modeled,
        "Residual": residual,
        "Relative Error": relative_error,
        "Uncertainty Flag": flags,
    })

    return output


def build_summary_table(fit: dict, table: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if fit.get("site"):
        rows.append({"Metric": "site", "Value": fit["site"]})
    rows.append({"Metric": "h0", "Value": fit["h0"]})

    if fit.get("is_segmented"):
        rows.append({"Metric": "breakpoint stage (m)", "Value": fit["breakpoint"]})
        for i, seg in enumerate(fit["segments"], start=1):
            rows.append({"Metric": f"segment {i} a", "Value": seg["a"]})
            rows.append({"Metric": f"segment {i} b", "Value": seg["b"]})
            rows.append({"Metric": f"segment {i} points", "Value": seg["n_points"]})
    else:
        rows.append({"Metric": "a", "Value": fit["a"]})
        rows.append({"Metric": "b", "Value": fit["b"]})

    rows += [
        {"Metric": "R^2", "Value": fit["r_squared"]},
        {"Metric": "Valid points", "Value": len(table)},
        {"Metric": "Uncertain points", "Value": int((table["Uncertainty Flag"] == "Uncertain").sum())},
        {"Metric": "Normal points", "Value": int((table["Uncertainty Flag"] == "Normal").sum())},
    ]
    return pd.DataFrame(rows)


def export_rating_curve_report(
    df: pd.DataFrame,
    output_path: str | Path,
    a: float,
    b: float,
    h0: float,
    uncertainty_threshold: float = 0.25,
    r_squared: float | None = None,
    fit: dict | None = None,
    site: str | None = None,
) -> Path:
    """Write the multi-sheet Excel report.

    Pass ``fit`` (the dict from :func:`fit_rating_curve`) to render segmented
    curves and reuse its overall R²; otherwise a single power law ``a``/``b``/
    ``h0`` is used. ``site`` (or a single-valued site column in ``df``) is
    recorded in the Summary sheet.
    """
    from src.rating_curve_fitting import predict_discharge

    predict = (lambda stage: predict_discharge(fit, stage)) if fit is not None else None

    if site is None and SITE in df.columns:
        unique_sites = df[SITE].dropna().astype(str).str.strip().unique()
        unique_sites = [s for s in unique_sites if s]
        if len(unique_sites) == 1:
            site = unique_sites[0]
    original_data = df.copy()
    table = build_observed_modeled_table(
        df, a=a, b=b, h0=h0, uncertainty_threshold=uncertainty_threshold, predict=predict
    )

    if fit is not None and r_squared is None:
        r_squared = fit.get("r_squared")

    if r_squared is None:
        if len(table) > 1:
            ss_res = float(np.sum((table[OUT_OBSERVED] - table[OUT_MODELED]) ** 2))
            ss_tot = float(np.sum((table[OUT_OBSERVED] - table[OUT_OBSERVED].mean()) ** 2))
            r_squared = 1.0 - (ss_res / ss_tot) if ss_tot != 0 else 1.0
        else:
            r_squared = 1.0

    fit_summary = dict(fit) if fit is not None else {"a": a, "b": b, "h0": h0}
    fit_summary["r_squared"] = r_squared
    if site:
        fit_summary["site"] = site
    summary = build_summary_table(fit_summary, table)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        original_data.to_excel(writer, sheet_name="Original Data", index=False)
        table.to_excel(writer, sheet_name="Observed vs Modeled", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

        plot_data = pd.DataFrame(
            {
                "Stage Above Bed (m)": table[OUT_STAGE],
                "Observed Discharge Q (m³/s)": table[OUT_OBSERVED],
                "Modeled Discharge Q (m³/s)": table[OUT_MODELED],
                "Uncertain Observed Discharge Q (m³/s)": np.where(table["Uncertainty Flag"] == "Uncertain", table[OUT_OBSERVED], np.nan),
            }
        )
        plot_data.to_excel(writer, sheet_name="Plot Data", index=False)

        workbook = writer.book
        plot_ws = workbook.create_sheet("Plot")
        plot_ws.append(["Stage Above Bed (m)", "Observed Discharge Q (m³/s)", "Modeled Discharge Q (m³/s)", "Uncertain Observed Discharge Q (m³/s)"])
        for row in plot_data.values.tolist():
            plot_ws.append(row)

        chart = LineChart()
        chart.title = "Rating Curve"
        chart.x_axis.title = "Stage Above Bed (m)"
        chart.y_axis.title = "Discharge (m³/s)"
        chart.style = 13
        chart.legend.position = "r"
        chart.height = 9
        chart.width = 18
        chart.y_axis.scaling.min = 0

        observed_data = Reference(plot_ws, min_col=2, max_col=2, min_row=1, max_row=plot_ws.max_row)
        modeled_data = Reference(plot_ws, min_col=3, max_col=3, min_row=1, max_row=plot_ws.max_row)
        uncertain_data = Reference(plot_ws, min_col=4, max_col=4, min_row=1, max_row=plot_ws.max_row)
        categories = Reference(plot_ws, min_col=1, max_col=1, min_row=2, max_row=plot_ws.max_row)

        chart.add_data(observed_data, titles_from_data=True)
        chart.add_data(modeled_data, titles_from_data=True)
        chart.add_data(uncertain_data, titles_from_data=True)
        chart.set_categories(categories)

        chart.series[0].graphicalProperties.line.solidFill = "1F77B4"
        chart.series[0].graphicalProperties.line.width = 2
        chart.series[0].marker.size = 6
        chart.series[0].marker.symbol = "circle"
        chart.series[0].marker.graphicalProperties.solidFill = "1F77B4"

        chart.series[1].graphicalProperties.line.solidFill = "2CA02C"
        chart.series[1].graphicalProperties.line.width = 2
        chart.series[1].marker.size = 4
        chart.series[1].marker.symbol = "triangle"
        chart.series[1].marker.graphicalProperties.solidFill = "2CA02C"

        chart.series[2].graphicalProperties.line.noFill = True
        chart.series[2].marker.size = 7
        chart.series[2].marker.symbol = "diamond"
        chart.series[2].marker.graphicalProperties.solidFill = "D62728"

        plot_ws.add_chart(chart, "F2")

        observed_ws = workbook["Observed vs Modeled"]
        flagged_fill = PatternFill(fill_type="solid", fgColor="FFC000")
        normal_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        for row_idx in range(2, observed_ws.max_row + 1):
            flag = observed_ws.cell(row=row_idx, column=7).value
            fill = flagged_fill if flag == "Uncertain" else normal_fill
            for col_idx in range(1, observed_ws.max_column + 1):
                observed_ws.cell(row=row_idx, column=col_idx).fill = fill

    return output


def main() -> None:
    import pandas as pd

    from src.rating_curve_fitting import fit_rating_curve

    input_path = Path(__file__).resolve().parent.parent / "cleaned_measurements.csv"
    output_path = Path(__file__).resolve().parent.parent / "rating_curve_report.xlsx"

    df = pd.read_csv(input_path)
    fit = fit_rating_curve(df)
    export_rating_curve_report(
        df, output_path, a=fit["a"], b=fit["b"], h0=fit["h0"], r_squared=fit["r_squared"], fit=fit
    )

    print(f"Rating-curve report written to: {output_path.name}")


if __name__ == "__main__":
    main()
